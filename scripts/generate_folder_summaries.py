from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SUMMARY_NAME = "FOLDER_SUMMARY.txt"
SKIP_DIR_NAMES = {
    ".git",
    ".venv",
    "__pycache__",
    "REINVENT4",
}
SKIP_PATH_PARTS = {
    "external_tools/SwissMADE",
}
MAX_FILES_LIST = 80
MAX_SCHEMA_FILES = 20


def should_skip_dir(path: Path) -> bool:
    path_str = path.as_posix()
    if any(part in path_str for part in SKIP_PATH_PARTS):
        return True
    return any(part in SKIP_DIR_NAMES for part in path.parts)


def purpose_for_folder(path: Path) -> str:
    rel = path.relative_to(ROOT).as_posix() if path != ROOT else "."
    rel_low = rel.lower()
    if rel == ".":
        return "Project root containing source code, generated results, configurations, documentation, and data assets."
    if rel_low.startswith("outputs/generated/mol2mol/plain"):
        return "Non-optimized Mol2Mol generation outputs scored with QSAR and ADME criteria."
    if rel_low.startswith("outputs/generated/mol2mol/reinforcement_learning"):
        return "Mol2Mol reinforcement-learning optimized outputs scored with QSAR and ADME criteria."
    if rel_low.startswith("outputs/generated/mol2mol/curriculum_learning"):
        return "Mol2Mol curriculum-learning optimized outputs scored with QSAR and ADME criteria."
    if rel_low.startswith("outputs/generated/reinforcement_learning"):
        return "REINVENT reinforcement-learning optimized molecule outputs and exports."
    if rel_low.startswith("outputs/generated/curriculum_learning"):
        return "REINVENT curriculum-learning optimized molecule outputs and exports."
    if rel_low.startswith("outputs/generated"):
        return "Generated molecule output exports, usually as CSV/XLSX with structure images and manifests."
    if rel_low.startswith("outputs/qsar"):
        return "QSAR prediction outputs and supporting exports."
    if rel_low.startswith("reinvent_integration/results/mol2mol_plain"):
        return "Raw Mol2Mol sampling outputs without optimization."
    if rel_low.startswith("reinvent_integration/results/mol2mol_rl_adme"):
        return "Mol2Mol reinforcement-learning training and sampling artifacts for QSAR+ADME optimization."
    if rel_low.startswith("reinvent_integration/results/mol2mol_cl_adme"):
        return "Mol2Mol curriculum-learning stage artifacts and sampling outputs for QSAR+ADME optimization."
    if rel_low.startswith("reinvent_integration/results/reinforcement_learning_adme"):
        return "REINVENT reinforcement-learning artifacts for QSAR+ADME optimization."
    if rel_low.startswith("reinvent_integration/results/curriculum_learning_adme"):
        return "REINVENT curriculum-learning artifacts for QSAR+ADME optimization."
    if rel_low.startswith("reinvent_integration/results"):
        return "Intermediate and final REINVENT training, sampling, and scoring artifacts."
    if rel_low.startswith("reinvent_integration/configs"):
        return "Configuration files for REINVENT generation, transfer learning, RL, curriculum learning, and sampling workflows."
    if rel_low.startswith("reinvent_integration/data"):
        return "Prepared input data used for REINVENT training or sampling."
    if rel_low.startswith("reinvent_integration/artifacts"):
        return "Saved QSAR model bundles and metadata used by generation workflows."
    if rel_low.startswith("pharm_outputs"):
        return "Pharmacophore pipeline outputs, hypotheses, screening results, and prepared structures."
    if rel_low.startswith("data"):
        return "Project data assets, prepared tables, or structure inputs."
    if rel_low.startswith("docs"):
        return "Documentation and project notes."
    if rel_low.startswith("notebooks"):
        return "Notebook assets and exported notebook-related materials."
    if rel_low.startswith("presentation"):
        return "Presentation source files and compiled support assets."
    if rel_low.startswith("results"):
        return "Legacy or general result files generated outside the organized outputs folders."
    if rel_low.startswith("catboost_info"):
        return "CatBoost training logs and metadata from QSAR model experiments."
    return "Project folder containing source, data, configuration, or generated artifacts."


def workflow_notes_for_folder(path: Path) -> list[str]:
    rel = path.relative_to(ROOT).as_posix() if path != ROOT else "."
    low = rel.lower()
    notes: list[str] = []
    if "mol2mol/plain" in low:
        notes.append("Used Mol2Mol prior sampling without RL or curriculum optimization.")
        notes.append("Generated molecules were post-scored with the saved QSAR model, adme-py, and admet_ai.")
    if "mol2mol/reinforcement_learning" in low or "mol2mol_rl_adme" in low:
        notes.append("Used Mol2Mol reinforcement learning with a combined QSAR + ADME objective.")
        notes.append("Optimization targeted predicted IC50 <= 5 uM, cLogP 1-3.5, logS >= -6.0, microsome clearance <= 70, hepatocyte clearance <= 80, and half-life >= 40.")
    if "mol2mol/curriculum_learning" in low or "mol2mol_cl_adme" in low:
        notes.append("Used Mol2Mol curriculum learning with a relaxed stage followed by a strict optimization stage.")
        notes.append("Final strict stage used the same optimality thresholds as the RL workflow.")
    if "/reinforcement_learning" in low and "mol2mol" not in low:
        notes.append("Used REINVENT reinforcement learning with the external QSAR + ADME scorer.")
    if "/curriculum_learning" in low and "mol2mol" not in low:
        notes.append("Used REINVENT curriculum learning with staged threshold tightening.")
    if low.startswith("outputs/generated"):
        notes.append("Excel outputs in this folder include a Structure image column for each SMILES row.")
    if low.startswith("reinvent_integration/configs"):
        notes.append("Files in this folder were used to define sampling, transfer learning, RL, curriculum learning, and model-family-specific runs.")
    if low.startswith("reinvent_integration/artifacts"):
        notes.append("The saved QSAR model bundle in this area is reused by downstream scoring scripts.")
    if low.startswith("pharm_outputs"):
        notes.append("This area stores pharmacophore preparation, hypothesis generation, and screening outputs.")
    return notes


def detect_csv_schema(path: Path) -> dict:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader, [])
    return {"type": "csv", "columns": header}


def detect_xlsx_schema(path: Path) -> dict:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            header = ["" if cell is None else str(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
            sheets.append({"sheet": ws.title, "columns": header})
        return {"type": "xlsx", "sheets": sheets}
    except Exception as exc:
        return {"type": "xlsx", "error": str(exc)}


def detect_schema(path: Path) -> dict | None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return detect_csv_schema(path)
    if suffix == ".xlsx":
        return detect_xlsx_schema(path)
    if suffix == ".json":
        try:
            with path.open("r", encoding="utf-8", errors="replace") as fh:
                data = json.load(fh)
            if isinstance(data, dict):
                return {"type": "json", "keys": list(data.keys())}
            if isinstance(data, list) and data and isinstance(data[0], dict):
                return {"type": "json", "list_item_keys": list(data[0].keys())}
            return {"type": "json", "summary": type(data).__name__}
        except Exception as exc:
            return {"type": "json", "error": str(exc)}
    return None


def file_summary_line(path: Path) -> str:
    size_kb = path.stat().st_size / 1024
    return f"- {path.name} ({size_kb:.1f} KB)"


def extract_manifest_rows(path: Path) -> list[str]:
    lines: list[str] = []
    try:
        with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                if not row:
                    continue
                parts = [f"{k}={v}" for k, v in row.items() if v not in (None, "")]
                if parts:
                    lines.append("- " + "; ".join(parts))
    except Exception:
        return []
    return lines


def write_summary_for_dir(path: Path) -> None:
    subdirs = sorted([p for p in path.iterdir() if p.is_dir() and not should_skip_dir(p)], key=lambda p: p.name.lower())
    files = sorted([p for p in path.iterdir() if p.is_file() and p.name != SUMMARY_NAME], key=lambda p: p.name.lower())

    lines: list[str] = []
    rel = path.relative_to(ROOT).as_posix() if path != ROOT else "."
    lines.append(f"Folder: {rel}")
    lines.append(f"Generated: {datetime.now().isoformat(timespec='seconds')}")
    lines.append("")
    lines.append("Purpose")
    lines.append("-------")
    lines.append(purpose_for_folder(path))
    lines.append("")

    notes = workflow_notes_for_folder(path)
    if notes:
        lines.append("What was done here")
        lines.append("------------------")
        for note in notes:
            lines.append(f"- {note}")
        lines.append("")

    lines.append("Contents")
    lines.append("--------")
    lines.append(f"Subfolders: {len(subdirs)}")
    for d in subdirs[:MAX_FILES_LIST]:
        lines.append(f"- {d.name}/")
    if len(subdirs) > MAX_FILES_LIST:
        lines.append(f"- ... and {len(subdirs) - MAX_FILES_LIST} more subfolders")
    lines.append(f"Files: {len(files)}")
    for f in files[:MAX_FILES_LIST]:
        lines.append(file_summary_line(f))
    if len(files) > MAX_FILES_LIST:
        lines.append(f"- ... and {len(files) - MAX_FILES_LIST} more files")
    lines.append("")

    schema_candidates = [f for f in files if f.suffix.lower() in {".csv", ".xlsx", ".json"}][:MAX_SCHEMA_FILES]
    if schema_candidates:
        lines.append("Schema / structure detected")
        lines.append("-------------------------")
        for f in schema_candidates:
            schema = detect_schema(f)
            lines.append(f"File: {f.name}")
            if schema is None:
                lines.append("- No schema extracted")
            elif schema.get("type") == "csv":
                lines.append("- Type: CSV")
                lines.append("- Columns: " + ", ".join(schema.get("columns", [])))
            elif schema.get("type") == "xlsx":
                lines.append("- Type: XLSX")
                if "error" in schema:
                    lines.append(f"- Schema read error: {schema['error']}")
                else:
                    for sheet in schema.get("sheets", []):
                        lines.append(f"- Sheet {sheet['sheet']}: " + ", ".join(sheet.get("columns", [])))
            elif schema.get("type") == "json":
                if "keys" in schema:
                    lines.append("- JSON keys: " + ", ".join(schema["keys"]))
                elif "list_item_keys" in schema:
                    lines.append("- JSON list item keys: " + ", ".join(schema["list_item_keys"]))
                else:
                    lines.append("- JSON summary: " + schema.get("summary", schema.get("error", "unknown")))
        lines.append("")

    manifests = [f for f in files if f.name.lower() == "manifest.csv"]
    if manifests:
        lines.append("Recorded output summary")
        lines.append("-----------------------")
        for manifest in manifests:
            manifest_lines = extract_manifest_rows(manifest)
            if manifest_lines:
                lines.extend(manifest_lines)
        lines.append("")

    lines.append("Other information")
    lines.append("-----------------")
    lines.append("- Summaries were generated automatically for project-owned folders.")
    lines.append("- Vendor, cache, and environment folders were skipped to avoid clutter.")
    lines.append("- Review manifests, logs, and scored CSV/XLSX files in this folder for detailed run-level results when present.")
    lines.append("")

    (path / SUMMARY_NAME).write_text("\n".join(lines), encoding="utf-8")


def iter_project_dirs(root: Path) -> Iterable[Path]:
    for path in [root] + sorted([p for p in root.rglob("*") if p.is_dir()], key=lambda p: p.as_posix()):
        if should_skip_dir(path):
            continue
        yield path


def main() -> int:
    count = 0
    for path in iter_project_dirs(ROOT):
        write_summary_for_dir(path)
        count += 1
    print(f"WROTE_SUMMARIES {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
