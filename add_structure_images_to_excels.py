from __future__ import annotations

import argparse
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from rdkit import Chem
from rdkit.Chem import AllChem, Draw
from rdkit import RDLogger

RDLogger.DisableLog('rdApp.*')

IMG_W = 170
IMG_H = 120
ROW_H_PT = 92
IMG_COL_W = 24
TARGET_GLOBS = ("**/*.xlsx",)
SKIP_DIR_PARTS = {".venv", "__pycache__", "external_tools/SwissMADE"}


def should_skip(path: Path) -> bool:
    path_str = str(path).replace('\\', '/')
    return any(part in path_str for part in SKIP_DIR_PARTS)


def smiles_headers(headers: list[str]) -> list[int]:
    hits = []
    for idx, header in enumerate(headers, start=1):
        h = (header or '').strip().lower()
        if 'smiles' in h and h != 'smiles_state':
            hits.append(idx)
    return hits


def pick_smiles_col(ws, candidate_cols: list[int]) -> int | None:
    if not candidate_cols:
        return None
    if len(candidate_cols) == 1:
        return candidate_cols[0]

    best_col = candidate_cols[0]
    best_score = -1
    max_probe = min(ws.max_row, 25)
    for col_idx in candidate_cols:
        score = 0
        for row_idx in range(2, max_probe + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if not value:
                continue
            mol = Chem.MolFromSmiles(str(value).strip())
            if mol is not None:
                score += 1
        if score > best_score:
            best_score = score
            best_col = col_idx
    return best_col


def add_structure_column_to_sheet(ws, temp_dir: Path) -> bool:
    headers = [str(cell.value) if cell.value is not None else '' for cell in ws[1]]
    if any(h.strip().lower() == 'structure' for h in headers):
        return False

    candidate_cols = smiles_headers(headers)
    smiles_col = pick_smiles_col(ws, candidate_cols)
    if smiles_col is None:
        return False

    ws.insert_cols(1)
    ws.cell(row=1, column=1, value='Structure')
    shifted_smiles_col = smiles_col + 1
    ws.column_dimensions['A'].width = IMG_COL_W

    created = 0
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = ROW_H_PT
        value = ws.cell(row=row_idx, column=shifted_smiles_col).value
        if value is None:
            continue
        smiles = str(value).strip()
        if not smiles:
            continue
        mol = Chem.MolFromSmiles(smiles)
        if mol is None:
            continue
        AllChem.Compute2DCoords(mol)
        img = Draw.MolToImage(mol, size=(IMG_W, IMG_H))
        png_path = temp_dir / f"{ws.title.replace('/', '_')}_{row_idx}.png"
        img.save(png_path)
        xl_img = XLImage(str(png_path))
        xl_img.width = IMG_W
        xl_img.height = IMG_H
        xl_img.anchor = f'A{row_idx}'
        ws.add_image(xl_img)
        created += 1

    return created > 0


def process_workbook(path: Path) -> bool:
    wb = load_workbook(path)
    changed = False
    with tempfile.TemporaryDirectory(prefix='excel_structures_') as tmp:
        temp_dir = Path(tmp)
        for ws in wb.worksheets:
            changed = add_structure_column_to_sheet(ws, temp_dir) or changed
        if changed:
            backup = path.with_suffix(path.suffix + '.bak')
            shutil.copy2(path, backup)
            wb.save(path)
    return changed


def find_workbooks(root: Path) -> list[Path]:
    files: list[Path] = []
    for pattern in TARGET_GLOBS:
        for path in root.glob(pattern):
            if path.is_file() and not should_skip(path):
                files.append(path)
    return sorted(set(files))


def main() -> int:
    parser = argparse.ArgumentParser(description='Add structure image columns to Excel workbooks with SMILES columns.')
    parser.add_argument('--root', default='.', help='Project root to scan')
    args = parser.parse_args()

    root = Path(args.root).resolve()
    updated: list[Path] = []
    for path in find_workbooks(root):
        try:
            if process_workbook(path):
                updated.append(path)
                print(f'UPDATED {path.relative_to(root)}')
        except Exception as exc:
            print(f'FAILED {path.relative_to(root)}: {exc}')

    print(f'UPDATED_COUNT {len(updated)}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
